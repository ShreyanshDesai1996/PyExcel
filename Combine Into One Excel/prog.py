import os
import openpyxl

# Get the current directory
directory = os.getcwd()

# Create a new workbook for the output
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Loop through the files in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(directory, filename)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Copy values from the first column and remove spaces
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                if row[0] is not None:
                    cell_value = str(row[0])
                    cell_value = cell_value.replace(" ", "")  # Remove spaces
                    cell_value = cell_value.replace(".0", "")  # Remove spaces
                    cell_value = cell_value.replace(",", "")  # Remove spaces
                    if len(cell_value) == 10 or len(cell_value) == 13:
                        output_ws.append([cell_value])

            wb.close()
        except FileNotFoundError:
            print(f"File {filename} not found.")

# Iterate over cells in the output workbook
for row_index, row in enumerate(output_ws.iter_rows(values_only=True), start=1):
    cell_value = str(row[0])
    if cell_value is not None and len(cell_value) == 10:
        new_value = "+91" + cell_value
        cell = output_ws.cell(row=row_index, column=1)
        cell.value = new_value

# Save the output workbook
output_wb.save("output.xlsx")
output_wb.close()
